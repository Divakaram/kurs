from django.db.models import Q
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from cms.models import *
from cms.forms import *
from crm.forms import *
from crm.models import Price as PriceView
from django.views.generic import ListView
from crm.telegram import *
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from datetime import timedelta
from openpyxl.styles import Font
from django.db.models import Count


def index(request):
    slider_list = CmsSlider.objects.all()
    napr = NaprCrm.objects.all()
    if request.method == 'POST':
        form = OrderForm(request.POST)
        name = request.POST['order_name']
        phone = request.POST['order_phone']
        if form.is_valid():
            send_message(tg_name=name, tg_phone=phone)
            form.save()
            return render(request, 'crm/thanks_page.html', {'slider_list': slider_list,'napr': napr, 'form': form, 'name': name})
        else:
            return render(request, 'crm/phone_error.html', {'slider_list': slider_list,'napr': napr, 'form': form})
    else:
        form = OrderForm()
    return render(request, 'crm/index.html',
                  {'slider_list': slider_list, 'form': form, 'napr': napr, 'title': 'Главная'})


def teachers(requst):
    form = OrderForm()
    posts = CmsTeachers.objects.all()
    return render(requst, 'crm/teachers.html', {'form': form, 'posts': posts, 'title': 'Тренеры'})


def price(request):
    form = OrderForm()
    price_site = PriceView.objects.all()
    return render(request, 'crm/price.html', {'price_site': price_site, 'form': form, 'title': 'Цены'})


@login_required
def dashboard(request):
    order = Order.objects.all()
    return render(request, 'crm/dashboard.html', {'order': order, 'title': "Панель управления"})


def show_post(request, post_slug):
    form = OrderForm()
    post = get_object_or_404(CmsTeachers, slug=post_slug)
    return render(request, 'crm/teacher_about.html', {'post': post, 'form': form, 'title': 'О тренерах'})


@login_required
def delete_order(request, id_order):
    order = Order.objects.get(pk=id_order)
    order.delete()
    return redirect('dashboard')


@login_required
def update_order(request, id_order):
    order = Order.objects.get(pk=id_order)
    form = OrderForm(request.POST or None, instance=order)
    if form.is_valid():
        form.save()
        return redirect('dashboard')
    return render(request, 'crm/update_order.html', {'order': order, 'form': form})


@login_required
def delete_teacher(request, id_teacher):
    teacher = CmsTeachers.objects.get(pk=id_teacher)
    teacher.delete()
    return redirect('admin_teachers')


@login_required
def update_teacher(request, id_teacher):
    teacher = CmsTeachers.objects.get(pk=id_teacher)
    form = TeacherForm(request.POST or None, request.FILES or None, instance=teacher)
    if form.is_valid():
        form.save()
        return redirect('admin_teachers')
    return render(request, 'crm/update_teacher.html', {'teacher': teacher, 'form': form})


@login_required
def add_teacher(request):
    form = TeacherForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        return redirect('admin_teachers')
    return render(request, 'crm/add_teacher.html', {'form': form})


@login_required
def delete_price(request, id_price):
    price = Price.objects.get(pk=id_price)
    price.delete()
    return redirect('admin_price')


@login_required
def update_price(request, id_price):
    price = Price.objects.get(pk=id_price)
    form = PriceForm(request.POST or None, instance=price)
    if form.is_valid():
        form.save()
        return redirect('admin_price')
    return render(request, 'crm/update_price.html', {'price': price, 'form': form})


@login_required
def add_price(request):
    form = PriceForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('admin_price')
    return render(request, 'crm/add_price.html', {'form': form})


@login_required
def delete_user(request, user_name):
    users = User.objects.get(username=user_name)
    users.delete()
    return redirect('show_users')


@login_required
def add_user(request):
    form = UserForm(request.POST or None)
    if form.is_valid():
        User.objects.create_user(**form.cleaned_data)
        return redirect('show_users')
    return render(request, 'crm/add_user.html', {'form': form})


@login_required
def admin_teachers(request):
    teacher = CmsTeachers.objects.all()
    return render(request, 'crm/admin_teachers.html', {'order': teacher, 'title': "Панель управления"})


@login_required
def admin_price(request):
    price = Price.objects.all()
    return render(request, 'crm/admin_price.html', {'order': price, 'title': "Панель управления"})


@login_required
def show_users(request):
    users = User.objects.all()
    return render(request, 'crm/admin_users.html', {'users': users, 'title': "Панель управления"})


@login_required
def admin_napr(request):
    napr = NaprCrm.objects.all()
    return render(request, 'crm/admin_napr.html', {'napr': napr, 'title': "Панель управления"})


@login_required
def delete_napr(request, id_napr):
    napr = NaprCrm.objects.get(pk=id_napr)
    napr.delete()
    return redirect('admin_napr')


@login_required
def update_napr(request, id_napr):
    napr = NaprCrm.objects.get(pk=id_napr)
    form = NaprForm(request.POST or None, request.FILES or None, instance=napr)
    if form.is_valid():
        form.save()
        return redirect('admin_napr')
    return render(request, 'crm/update_napr.html', {'napr': napr, 'form': form})


@login_required
def add_napr(request):
    form = NaprForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        return redirect('admin_napr')
    return render(request, 'crm/add_napr.html', {'form': form})


def export_to_xlsx(request):
    order = Order.objects.filter(date__month=datetime.now().month)
    count = order.count()
    popular_napr = Order.objects.filter(date__month=datetime.now().month).values('order_napr__napr_name').annotate(
        total=Count("order_napr")).order_by("-total")[:1]
    popular_napr_value = popular_napr[0]
    p = popular_napr_value['order_napr__napr_name']
    popular_price = Order.objects.filter(date__month=datetime.now().month).values('order_price__price_name').annotate(
        total=Count("order_price")).order_by("-total")[:1]
    popular_price_value = popular_price[0]
    p_price = popular_price_value['order_price__price_name']
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}_otchet.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Отчёт'
    columns = [
        'Имя',
        'Телефон',
        'Направление',
        'Абонемент',
        'Дата',
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for ord in order:
        row_num += 1
        row = [
            ord.order_name,
            ord.order_phone,
            ord.order_napr.napr_name,
            ord.order_price.price_name,
            ord.date,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    cell_count_name = worksheet.cell(row=1, column=7)
    cell_count_name.value = "Количество заявок за месяц"
    cell_count_name.font = Font(name='Calibri', bold=True, size=14)
    column_letter = get_column_letter(7)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 35
    column_letter = get_column_letter(8)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 21

    cell_count_value = worksheet.cell(row=1, column=8)
    cell_count_value.value = count
    cell_count_value.font = Font(size=14)

    cell_count_name = worksheet.cell(row=2, column=7)
    cell_count_name.value = "Популярное направление"
    cell_count_name.font = Font(name='Calibri', bold=True, size=14)

    cell_count_value = worksheet.cell(row=2, column=8)
    cell_count_value.value = p
    cell_count_value.font = Font(size=14)

    cell_count_name = worksheet.cell(row=3, column=7)
    cell_count_name.value = "Популярный абонемент"
    cell_count_name.font = Font(name='Calibri', bold=True, size=14)

    cell_count_value = worksheet.cell(row=3, column=8)
    cell_count_value.value = p_price
    cell_count_value.font = Font(size=14)

    for i in range(2, len(columns) + 1):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    workbook.save(response)
    return response
